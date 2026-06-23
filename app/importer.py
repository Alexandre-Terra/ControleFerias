"""Importação idempotente da planilha Excel para o banco.

Lida com a estrutura multi-linha das abas de empresa: a primeira linha de um
funcionário tem código (A) + nome (C); linhas seguintes sem código mas com
início aquisitivo (Q) são períodos adicionais do mesmo funcionário.

A planilha é base apenas para funcionários e períodos aquisitivos —
**programações de férias nascem exclusivamente no app** e nunca são
importadas (as colunas W/X da planilha são ignoradas).

Modo ``dry_run``: faz tudo dentro de uma transação e dá rollback no final,
permitindo conferir contagens (novos / atualizados / inalterados) e divergências
de consistência antes de gravar de verdade.
"""
import unicodedata
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

from .models import Empresa, Funcionario, PeriodoAquisitivo, Setor, db

EXCEL_EPOCH = date(1899, 12, 30)

ABAS_IGNORADAS = {"DASHBOARD", "Resumo"}
PRIMEIRA_LINHA_DADOS = 8

# Índices de coluna (1-based) conforme o cabeçalho das abas de empresa.
COL = {
    "codigo": 1,    # A
    "setor": 2,     # B - setor (texto livre; só na linha do funcionário)
    "nome": 3,      # C
    "admissao": 10, # J
    "vencto": 11,   # K
    "q_inicio": 17, # Q - início aquisitivo
    "r_fim": 18,    # R - fim aquisitivo
    "z_abono": 26,  # Z - abono
    "ab_13": 28,    # AB - 13º
    "ac_direito": 29,   # AC - dias de direito
    "ag_restante": 33,  # AG - dias restantes
    "ah_limite": 34,    # AH - limite p/ gozo
}

CAMPOS_FUNC = ("nome", "data_admissao", "vencto_ferias")
CAMPOS_PERIODO = (
    "fim", "dias_direito", "dias_restantes", "limite_gozo",
    "dias_abono", "decimo_terceiro",
)


def to_date(v):
    """Converte serial Excel / datetime / string em date, ou None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        return EXCEL_EPOCH + timedelta(days=int(v)) if v > 0 else None
    s = str(v).strip()
    if not s or set(s) <= set("./ -"):  # ex.: "..../..../......"
        return None
    try:
        return EXCEL_EPOCH + timedelta(days=int(float(s.replace(",", "."))))
    except ValueError:
        return None


def to_float(v):
    """Converte número / string (vírgula ou ponto) em float, ou None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if not s or set(s) <= set("./ -"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _texto(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# A coluna B traz o setor como texto livre minúsculo, às vezes com acento
# faltando ("licitaçao"). Mapeamos para um vocabulário canônico único: como
# `setor.nome` é UNIQUE e case/acento-sensitive, gravar "oficina" criaria um
# segundo setor além do "Oficina" já semeado (seeds/setores.py) e dividiria os
# dados. A chave do mapa é sem acento e minúscula, então variações casam.
SETOR_CANONICO = {
    "oficina": "Oficina",
    "comercial": "Comercial",
    "logistica": "Logística",
    "licitacao": "Licitação",
    "obras": "Obras",
    "adm": "Administrativo",
}


def _sem_acento(s):
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def setor_canonico(raw):
    """Texto livre da coluna B -> (nome canônico, conhecido).

    Retorna ``(None, False)`` para célula vazia. Casa por chave sem acento e
    minúscula ("LOGISTICA", "logística", "licitaçao" → mesmo setor). Valor fora
    do mapa vira Title-case e ``conhecido=False`` para o chamador sinalizar.
    """
    txt = _texto(raw)
    if txt is None:
        return None, False
    chave = _sem_acento(txt).lower()
    if chave in SETOR_CANONICO:
        return SETOR_CANONICO[chave], True
    return txt.capitalize(), False


def _get_or_create(model, defaults=None, **filtros):
    obj = db.session.query(model).filter_by(**filtros).first()
    if obj is None:
        obj = model(**filtros, **(defaults or {}))
        db.session.add(obj)
        criado = True
    else:
        criado = False
    return obj, criado


def _snapshot(obj, campos):
    return tuple(getattr(obj, c) for c in campos)


def _validar_periodo(p, ref, avisos):
    if p.fim and p.inicio:
        delta = (p.fim - p.inicio).days
        if not 300 <= delta <= 400:
            avisos.append(
                f"{ref}: período aquisitivo com {delta} dias "
                f"({p.inicio} → {p.fim}); esperado ~365"
            )
    if p.fim and p.limite_gozo:
        delta = (p.limite_gozo - p.fim).days
        if delta < 0:
            avisos.append(
                f"{ref}: limite_gozo {p.limite_gozo} anterior ao fim {p.fim}"
            )
        elif not 300 <= delta <= 400:
            avisos.append(
                f"{ref}: limite_gozo {delta} dias após o fim ({p.fim} → "
                f"{p.limite_gozo}); esperado ~365"
            )
    if p.dias_direito is not None and p.dias_restantes is not None:
        if p.dias_restantes > p.dias_direito:
            avisos.append(
                f"{ref}: dias_restantes ({p.dias_restantes}) > "
                f"dias_direito ({p.dias_direito})"
            )


def importar_xlsx(caminho, dry_run=False):
    """Importa a planilha. Retorna relatório estruturado.

    Quando ``dry_run=True``, dá rollback no final em vez de commit — os
    contadores e avisos refletem o que *seria* gravado.
    """
    wb = load_workbook(caminho, data_only=True, read_only=True)

    contadores = {
        e: {"novos": 0, "atualizados": 0, "inalterados": 0}
        for e in ("empresas", "funcionarios", "periodos")
    }
    avisos = []

    def registrar(entidade, novo, antes, depois):
        if novo:
            contadores[entidade]["novos"] += 1
        elif antes != depois:
            contadores[entidade]["atualizados"] += 1
        else:
            contadores[entidade]["inalterados"] += 1

    for ws in wb.worksheets:
        if ws.title in ABAS_IGNORADAS:
            continue

        empresa, criada = _get_or_create(Empresa, nome=ws.title)
        if criada:
            db.session.flush()
        # Empresa não tem campos atualizáveis pelo importer.
        registrar("empresas", criada, (), ())

        funcionario_atual = None

        for linha_idx, row in enumerate(
            ws.iter_rows(min_row=PRIMEIRA_LINHA_DADOS, values_only=True),
            start=PRIMEIRA_LINHA_DADOS,
        ):
            def cell(key):
                idx = COL[key] - 1
                return row[idx] if idx < len(row) else None

            ref = f"{ws.title}:L{linha_idx}"
            codigo = _texto(cell("codigo"))
            q_inicio = to_date(cell("q_inicio"))

            if codigo:
                nome = _texto(cell("nome")) or "(sem nome)"
                funcionario_atual, novo = _get_or_create(
                    Funcionario, empresa_id=empresa.id, codigo=codigo
                )
                antes = _snapshot(funcionario_atual, CAMPOS_FUNC)
                funcionario_atual.nome = nome
                funcionario_atual.data_admissao = to_date(cell("admissao"))
                funcionario_atual.vencto_ferias = to_date(cell("vencto"))
                db.session.flush()
                depois = _snapshot(funcionario_atual, CAMPOS_FUNC)
                registrar("funcionarios", novo, antes, depois)
            elif q_inicio is None:
                continue  # linha em branco / separador

            if q_inicio is None or funcionario_atual is None:
                continue

            periodo, novo_p = _get_or_create(
                PeriodoAquisitivo,
                funcionario_id=funcionario_atual.id,
                inicio=q_inicio,
            )
            antes_p = _snapshot(periodo, CAMPOS_PERIODO)
            periodo.fim = to_date(cell("r_fim"))
            periodo.dias_direito = to_float(cell("ac_direito"))
            periodo.dias_restantes = to_float(cell("ag_restante"))
            periodo.limite_gozo = to_date(cell("ah_limite"))
            periodo.dias_abono = to_float(cell("z_abono"))
            periodo.decimo_terceiro = _texto(cell("ab_13"))
            db.session.flush()
            depois_p = _snapshot(periodo, CAMPOS_PERIODO)
            registrar("periodos", novo_p, antes_p, depois_p)
            _validar_periodo(periodo, ref, avisos)

    if dry_run:
        db.session.rollback()
    else:
        db.session.commit()
    wb.close()

    return {**contadores, "avisos": avisos, "dry_run": dry_run}


def importar_setores(caminho, dry_run=False):
    """Atribui setores aos funcionários a partir da coluna B da planilha.

    Casa cada funcionário por **empresa (título da aba) + código** — os
    funcionários já devem existir (rode ``import-xlsx`` antes). Cria os setores
    faltantes a partir do vocabulário canônico (``SETOR_CANONICO``), evitando
    duplicar os já semeados. Idempotente; ``dry_run=True`` dá rollback no final.

    Retorna contadores, avisos e ``nao_encontrados`` (códigos da planilha sem
    funcionário correspondente no banco — sintoma de prod desatualizada).
    """
    wb = load_workbook(caminho, data_only=True, read_only=True)

    contadores = {
        "setores_criados": 0, "atribuidos": 0, "inalterados": 0, "sem_setor": 0,
    }
    avisos = []
    nao_encontrados = []

    for ws in wb.worksheets:
        if ws.title in ABAS_IGNORADAS:
            continue

        empresa = db.session.query(Empresa).filter_by(nome=ws.title).first()
        if empresa is None:
            avisos.append(
                f"empresa '{ws.title}' não existe no banco; aba ignorada"
            )
            continue

        for linha_idx, row in enumerate(
            ws.iter_rows(min_row=PRIMEIRA_LINHA_DADOS, values_only=True),
            start=PRIMEIRA_LINHA_DADOS,
        ):
            def cell(key):
                idx = COL[key] - 1
                return row[idx] if idx < len(row) else None

            codigo = _texto(cell("codigo"))
            if not codigo:
                continue  # linha de continuação (período extra) / separador

            ref = f"{ws.title}:L{linha_idx}"
            nome_setor, conhecido = setor_canonico(cell("setor"))
            if nome_setor is None:
                contadores["sem_setor"] += 1
                continue
            if not conhecido:
                avisos.append(
                    f"{ref}: setor '{_texto(cell('setor'))}' fora do mapa "
                    f"canônico; usando '{nome_setor}'"
                )

            setor, criado = _get_or_create(Setor, nome=nome_setor)
            if criado:
                db.session.flush()
                contadores["setores_criados"] += 1

            func = (
                db.session.query(Funcionario)
                .filter_by(empresa_id=empresa.id, codigo=codigo)
                .first()
            )
            if func is None:
                nao_encontrados.append(
                    f"{ref}: código {codigo} sem funcionário no banco"
                )
                continue

            if func.setor_id == setor.id:
                contadores["inalterados"] += 1
            else:
                func.setor_id = setor.id
                contadores["atribuidos"] += 1

    db.session.flush()
    if dry_run:
        db.session.rollback()
    else:
        db.session.commit()
    wb.close()

    return {
        **contadores,
        "avisos": avisos,
        "nao_encontrados": nao_encontrados,
        "dry_run": dry_run,
    }
