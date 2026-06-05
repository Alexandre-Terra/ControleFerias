"""Testes da importação de setores (coluna B): canonicalização, atribuição,
idempotência, dry-run e funcionário ausente."""
from openpyxl import Workbook

from app.importer import COL, importar_setores, setor_canonico
from app.models import Empresa, Funcionario, Setor, db


def test_setor_canonico_normaliza_acento_e_caixa():
    assert setor_canonico("oficina") == ("Oficina", True)
    assert setor_canonico("LOGISTICA") == ("Logística", True)
    assert setor_canonico(" licitaçao ") == ("Licitação", True)  # acento faltando
    assert setor_canonico("adm") == ("Administrativo", True)
    assert setor_canonico("Compras") == ("Compras", False)       # fora do mapa
    assert setor_canonico(None) == (None, False)
    assert setor_canonico("   ") == (None, False)


def _seed_funcs(empresa_nome, codigos):
    emp = Empresa(nome=empresa_nome)
    db.session.add(emp)
    db.session.flush()
    for cod in codigos:
        db.session.add(Funcionario(empresa_id=emp.id, codigo=cod, nome=f"F{cod}"))
    db.session.commit()
    return emp


def _planilha(tmp_path):
    """Aba 'Auto Mec': código 3→oficina, 92→comercial; uma linha de continuação
    (sem código) com setor que deve ser ignorada."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Auto Mec"

    def put(row, key, value):
        ws.cell(row=row, column=COL[key], value=value)

    put(8, "codigo", 3)
    put(8, "nome", "DOUGLAS")
    put(8, "setor", "oficina")
    # Linha 9: continuação (sem código) — setor preenchido NÃO deve vazar.
    put(9, "setor", "logistica")
    put(10, "codigo", 92)
    put(10, "nome", "LUANA")
    put(10, "setor", "comercial")

    caminho = tmp_path / "setores.xlsx"
    wb.save(caminho)
    return caminho


def test_atribui_setores_e_cria_faltantes(app, tmp_path):
    _seed_funcs("Auto Mec", ["3", "92"])
    caminho = _planilha(tmp_path)

    rel = importar_setores(str(caminho))

    assert rel["dry_run"] is False
    assert rel["atribuidos"] == 2
    assert rel["nao_encontrados"] == []
    assert rel["avisos"] == []

    assert Funcionario.query.filter_by(codigo="3").one().setor.nome == "Oficina"
    assert Funcionario.query.filter_by(codigo="92").one().setor.nome == "Comercial"

    nomes = {s.nome for s in Setor.query.all()}
    assert {"Oficina", "Comercial"} <= nomes
    # Linha de continuação (sem código) não criou "Logística".
    assert "Logística" not in nomes


def test_idempotente(app, tmp_path):
    _seed_funcs("Auto Mec", ["3", "92"])
    caminho = _planilha(tmp_path)

    importar_setores(str(caminho))
    rel = importar_setores(str(caminho))  # segunda vez não muda nada

    assert rel["atribuidos"] == 0
    assert rel["inalterados"] == 2
    assert rel["setores_criados"] == 0


def test_dry_run_nao_persiste(app, tmp_path):
    _seed_funcs("Auto Mec", ["3", "92"])
    caminho = _planilha(tmp_path)

    rel = importar_setores(str(caminho), dry_run=True)

    assert rel["dry_run"] is True
    assert rel["atribuidos"] == 2
    # Nada gravado — rollback descartou setor criado e atribuição.
    assert Funcionario.query.filter_by(codigo="3").one().setor_id is None
    assert Setor.query.filter_by(nome="Oficina").first() is None


def test_funcionario_ausente_vira_aviso(app, tmp_path):
    _seed_funcs("Auto Mec", ["3"])  # 92 não existe no banco
    caminho = _planilha(tmp_path)

    rel = importar_setores(str(caminho))

    assert rel["atribuidos"] == 1
    assert len(rel["nao_encontrados"]) == 1
    assert "92" in rel["nao_encontrados"][0]


def test_setor_desconhecido_atribui_e_avisa(app, tmp_path):
    _seed_funcs("Auto Mec", ["5"])
    wb = Workbook()
    ws = wb.active
    ws.title = "Auto Mec"
    ws.cell(row=8, column=COL["codigo"], value=5)
    ws.cell(row=8, column=COL["nome"], value="ZE")
    ws.cell(row=8, column=COL["setor"], value="compras")
    caminho = tmp_path / "desconhecido.xlsx"
    wb.save(caminho)

    rel = importar_setores(str(caminho))

    assert rel["atribuidos"] == 1
    assert Funcionario.query.filter_by(codigo="5").one().setor.nome == "Compras"
    assert any("fora do mapa" in a for a in rel["avisos"])


def test_empresa_inexistente_vira_aviso(app, tmp_path):
    # Nenhuma empresa "Auto Mec" no banco.
    caminho = _planilha(tmp_path)

    rel = importar_setores(str(caminho))

    assert rel["atribuidos"] == 0
    assert any("não existe no banco" in a for a in rel["avisos"])
