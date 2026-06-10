"""Testes do importador: conversões, parsing multi-linha, idempotência e dry-run."""
from datetime import date

from openpyxl import Workbook

from app.importer import (
    COL,
    EXCEL_EPOCH,
    importar_xlsx,
    to_date,
    to_float,
)
from app.models import Funcionario, PeriodoAquisitivo, ProgramacaoFerias, db


def serial(d):
    return (d - EXCEL_EPOCH).days


def test_to_date_variantes():
    assert to_date(serial(date(2024, 8, 1))) == date(2024, 8, 1)
    assert to_date(date(2024, 8, 1)) == date(2024, 8, 1)
    assert to_date("..../..../......") is None
    assert to_date(None) is None


def test_to_float_virgula_e_ponto():
    assert to_float("7,5") == 7.5
    assert to_float("30") == 30.0
    assert to_float(22.5) == 22.5
    assert to_float("-") is None
    assert to_float("") is None


def _planilha(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Auto Mec"  # aba de empresa (não ignorada)

    def put(row, key, value):
        ws.cell(row=row, column=COL[key], value=value)

    # Linha 8: funcionário com 1º período (fechado) + programação (W).
    put(8, "codigo", 3)
    put(8, "nome", "DOUGLAS SOUSA DA SILVA")
    put(8, "admissao", serial(date(2015, 10, 1)))
    put(8, "q_inicio", serial(date(2024, 10, 1)))
    put(8, "r_fim", serial(date(2025, 9, 30)))
    put(8, "ac_direito", 30)
    put(8, "ag_restante", 16)
    put(8, "ah_limite", serial(date(2026, 9, 15)))
    put(8, "w_gozo", serial(date(2026, 7, 1)))
    put(8, "x_dias", 16)

    # Linha 9: continuação (SEM código) -> 2º período do mesmo funcionário.
    put(9, "q_inicio", serial(date(2025, 10, 1)))
    put(9, "r_fim", serial(date(2026, 9, 30)))
    put(9, "ac_direito", "7,5")  # decimal com vírgula
    put(9, "ag_restante", "7,5")

    caminho = tmp_path / "mini.xlsx"
    wb.save(caminho)
    return caminho


def test_import_multilinha_e_conversoes(app, tmp_path):
    caminho = _planilha(tmp_path)
    rel = importar_xlsx(str(caminho))

    assert rel["empresas"]["novos"] == 1
    assert rel["funcionarios"]["novos"] == 1
    assert rel["periodos"]["novos"] == 2
    assert rel["programacoes"]["novos"] == 1
    assert rel["dry_run"] is False
    assert rel["avisos"] == []

    f = Funcionario.query.filter_by(codigo="3").one()
    assert f.nome == "DOUGLAS SOUSA DA SILVA"
    assert f.data_admissao == date(2015, 10, 1)
    assert len(f.periodos) == 2

    p2 = PeriodoAquisitivo.query.filter_by(inicio=date(2025, 10, 1)).one()
    assert p2.dias_direito == 7.5  # vírgula convertida

    prog = ProgramacaoFerias.query.one()
    assert prog.data_inicio == date(2026, 7, 1)
    assert prog.data_fim == date(2026, 7, 16)  # início + 16 - 1
    assert prog.origem == "import"


def test_import_idempotente(app, tmp_path):
    caminho = _planilha(tmp_path)
    importar_xlsx(str(caminho))
    rel = importar_xlsx(str(caminho))  # segunda vez não cria/atualiza nada

    for ent, esperado_inalterado in [
        ("empresas", 1),
        ("funcionarios", 1),
        ("periodos", 2),
        ("programacoes", 1),
    ]:
        assert rel[ent]["novos"] == 0, ent
        assert rel[ent]["atualizados"] == 0, ent
        assert rel[ent]["inalterados"] == esperado_inalterado, ent

    assert Funcionario.query.count() == 1
    assert PeriodoAquisitivo.query.count() == 2


def test_dry_run_nao_persiste(app, tmp_path):
    caminho = _planilha(tmp_path)
    rel = importar_xlsx(str(caminho), dry_run=True)

    assert rel["dry_run"] is True
    assert rel["funcionarios"]["novos"] == 1
    assert rel["periodos"]["novos"] == 2

    # Nada foi gravado — rollback descartou tudo.
    assert Funcionario.query.count() == 0
    assert PeriodoAquisitivo.query.count() == 0
    assert ProgramacaoFerias.query.count() == 0


def test_avisos_de_divergencia(app, tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Auto Mec"

    def put(row, key, value):
        ws.cell(row=row, column=COL[key], value=value)

    # Linha 8: dias_restantes > dias_direito + gozo após limite_gozo.
    put(8, "codigo", 7)
    put(8, "nome", "FULANO DA SILVA")
    put(8, "q_inicio", serial(date(2023, 1, 1)))
    put(8, "r_fim", serial(date(2023, 12, 31)))
    put(8, "ac_direito", 30)
    put(8, "ag_restante", 40)            # > dias_direito → aviso
    put(8, "ah_limite", serial(date(2024, 12, 30)))
    put(8, "w_gozo", serial(date(2025, 6, 1)))  # depois do limite → aviso
    put(8, "x_dias", 10)

    caminho = tmp_path / "div.xlsx"
    wb.save(caminho)

    rel = importar_xlsx(str(caminho), dry_run=True)

    assert len(rel["avisos"]) == 2
    assert any("dias_restantes" in a for a in rel["avisos"])
    assert any("após o limite" in a for a in rel["avisos"])


def test_aviso_programacao_import_orfa(app, tmp_path):
    caminho = _planilha(tmp_path)
    importar_xlsx(str(caminho))

    # Programação import sem linha correspondente na planilha (ex.: W/X
    # preenchidos por engano e depois corrigidos na origem).
    f = Funcionario.query.filter_by(codigo="3").one()
    db.session.add(ProgramacaoFerias(
        funcionario_id=f.id,
        data_inicio=date(2026, 6, 1),
        dias_gozo=30,
        data_fim=date(2026, 6, 30),
        origem="import",
    ))
    db.session.commit()

    rel = importar_xlsx(str(caminho), dry_run=True)
    orfas = [a for a in rel["avisos"] if "sem correspondência" in a]
    assert len(orfas) == 1
    assert "2026-06-01" in orfas[0]
